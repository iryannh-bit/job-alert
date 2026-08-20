"""
Charge la configuration depuis le fichier Excel (local ou Google Drive).

Sheets lus :
  - "Pour plus tard"      → config["pour_plus_tard"]
  - "Recherche entreprise" → config["recherche"]

Pour utiliser Google Drive :
  - Partage le fichier Excel en "Toute personne avec le lien peut voir"
  - Copie l'ID du fichier (dans l'URL Drive) dans GOOGLE_DRIVE_FILE_ID
  - Le script télécharge automatiquement la dernière version à chaque run

Pour utiliser un fichier local :
  - Mets le fichier Excel dans le repo (ex: suivi_stages.xlsx)
  - Laisse GOOGLE_DRIVE_FILE_ID vide
"""
import os
import io
import logging
import requests
import openpyxl

logger = logging.getLogger("config_loader")

LOCAL_EXCEL_PATH   = os.environ.get("EXCEL_PATH", "suivi_stages.xlsx")
GDRIVE_FILE_ID     = os.environ.get("GOOGLE_DRIVE_FILE_ID", "")

SHEET_PPT   = "Pour plus tard"
SHEET_RECH  = "Recherche entreprise"


def load_config() -> dict:
    """Charge et retourne la config depuis l'Excel."""
    wb = _load_workbook()
    return {
        "pour_plus_tard": _parse_pour_plus_tard(wb[SHEET_PPT]),
        "recherche":      _parse_recherche(wb[SHEET_RECH]),
    }


def _load_workbook() -> openpyxl.Workbook:
    if GDRIVE_FILE_ID:
        logger.info(f"Téléchargement Excel depuis Google Drive (id={GDRIVE_FILE_ID})…")
        url = f"https://drive.google.com/uc?export=download&id={GDRIVE_FILE_ID}"
        r = requests.get(url, timeout=30)
        r.raise_for_status()
        return openpyxl.load_workbook(io.BytesIO(r.content))
    else:
        logger.info(f"Chargement Excel local : {LOCAL_EXCEL_PATH}")
        return openpyxl.load_workbook(LOCAL_EXCEL_PATH)


def _parse_pour_plus_tard(ws) -> list[dict]:
    """
    Colonnes attendues (ligne 3 = en-têtes, données à partir de la ligne 4) :
    A: Contact  B: Entreprise  C: Poste  D: Infos complémentaires  E: Date de commencement
    """
    entries = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        contact, company, position, info, date_range = (row + (None,) * 5)[:5]
        if not company:
            continue
        entries.append({
            "contact":    str(contact or "").strip(),
            "company":    str(company).strip(),
            "position":   str(position or "").strip(),
            "info":       str(info or "").strip(),
            "date_range": str(date_range or "").strip(),
        })
    logger.debug(f"PPT : {len(entries)} entrées chargées.")
    return entries


def _parse_recherche(ws) -> list[dict]:
    """
    Colonnes attendues (ligne 3 = en-têtes, données à partir de la ligne 4) :
    A: Type de poste  B: Type d'entreprise  C: Sites  D: Infos  E: Date de début  F: Localisation
    """
    entries = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        job_type, company_type, sites, info, date_range, location = (row + (None,) * 6)[:6]
        if not job_type:
            continue
        entries.append({
            "job_type":     str(job_type).strip(),
            "company_type": str(company_type or "").strip(),
            "sites":        [s.strip() for s in str(sites or "").split(",")],
            "info":         str(info or "").strip(),
            "date_range":   str(date_range or "").strip(),
            "location":     str(location or "").strip(),
        })
    logger.debug(f"Recherche : {len(entries)} entrées chargées.")
    return entries
